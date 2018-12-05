from openpyxl import load_workbook,Workbook
from openpyxl.styles import Border,Side,Font
import time
from Util.var import  *


class parseExcel(object):

    def __init__(self,excelpath):
        self.excelpath = excelpath
        self.workbook = load_workbook(excelpath)  #加载excel
        self.sheet =  self.workbook.active #获取第一个sheet
        self.font = Font(color=None)
        self.colorDict = {'red','FFFF3030','green','FF008b00'}

    #设置当前操作的sheet对象,通过指数index,指定sheet
    def set_sheet_by_index(self,sheet_ineex):
        sheet_name =  self.workbook.sheetnames[sheet_ineex]
        self.sheet =  self.workbook[sheet_name]
        return self.sheet

    #获取sheet名称
    def get_sheet_name(self):
        return self.sheet.title

    #设置当前要操作的sheet对象,使用sheet名称获取相应的sheet
    def set_sheet_by_name(self,sheet_name):
        self.sheet  =  self.workbook[sheet_name]
        return self.sheet

    #获取默认sheet最大行数
    def get_sheet_max_row(self):
        return self.sheet.max_row

    #获取默认最大列
    def get_sheet_max_column(self):
        return self.sheet.max_column

    #获取sheet默认最小起始行号
    def get_min_row_no(self):
        return self.sheet.min_row

    #获取sheet默认最小起始列好
    def get_min_column_no(self):
        return self.sheet.min_column

    #获取sheet所有行对象
    def get_all_rows(self):
        return list(self.sheet.iter_rows())

    #获取sheet所有列对象
    def get_all_cols(self):
        return  list(self.sheet.iter_cols())

    #从默认的sheet中获取某一列，第一列从0开始
    def  get_single_col(self,col_no):
        return self.get_all_cols()[col_no]

    #从默认的sheet中获取某一行，第一行从0开始
    def get_singel_row(self,row_no):
        return  self.get_all_rows()[row_no]

    #从默认sheet中，通过提供的行号和列号，获取单元格,注意单元格行号和列从1开始
    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row_no,col_no)


    #从默认sheet中，通过提供的行号和列号，获取单元格,注意单元格行号和列从1开始
    def get_cell_content(self,row_no,col_no):
        return self.sheet.cell(row_no,col_no).value

    #向指定单元格输入内容
    def write_cell_content(self,row_no,col_no,content):
        self.sheet.cell(row_no,col_no).value =  content
        self.workbook.save(self.excelpath)
        return self.sheet.cell(row_no,col_no).value

    #向指定单元格输入时间
    def write_cell_content_time(self,row_no,col_no):
        time1 = time.strftime('%b %d %Y %H:%M:%S')
        self.sheet.cell(row_no, col_no).value = str(time1)
        self.workbook.save(self.excelpath)
        return self.sheet.cell(row_no, col_no).value

    #保存文件
    def save_ecxel_file(self):
        self.workbook.save(self.excelpath)

if __name__ == '__main__':
   #测试单元
   p = parseExcel(Excelobject_path)
   print('获取默认sheet:',p.get_sheet_name())
   print('获取sheet索引为1',p.set_sheet_by_index(1))
   print('获取默认sheet:',p.get_sheet_name())
   print('设置sheet索引为0',p.set_sheet_by_index(0))
   print('获取默认shet',p.get_sheet_name())
   print('获取最大行数',p.get_sheet_max_row())
   print('获取最大列',p.get_sheet_max_column())
   print('获取最小起始行数',p.get_min_row_no())
   print('获取最小起始列数',p.get_min_column_no())
   print('获取所有行的对象',p.get_all_rows())
   print('获取所有列的对象',p.get_all_cols())
   print('获取某一列',p.get_single_col(2))
   print('获取某一行',p.get_singel_row(2))
   print('获取行号和列号(2,2)单元格',p.get_cell(2,2))
   print('获取行号和列号(2,2)单元格内容呢',p.get_cell_content(2,2))
   print('通过行号和列号写入内容',p.write_cell_content(2,2,'邮箱地址'))
   print('通过行号和列号写入时间',p.write_cell_content_time(2,7))